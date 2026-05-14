"""
CCASS Inter-Broker Movement Analysis
=====================================
对比两个交易日，找出任意经纪商（Broker）持股变化超过 ±30% 的股票。

逻辑：
  1. 从 stocklist.aspx 获取全量股票列表（JSON API）
  2. 对每只股票，分别查询两个日期的持股数据
  3. 筛选 Participant ID 以 "B" 开头的行（经纪商）
  4. 对比同一经纪商在两日的持股，变化率超过阈值则记录该股票

使用方法：
    python q5.py

依赖：
    pip install requests beautifulsoup4 pandas openpyxl

修复说明（2026-05）：
  Bug 1 - STOCKLIST_URL：
    旧值指向 ccass_stock_list.htm（纯前端 SPA，直接请求返回空 HTML，resp.json() 报错）。
    真正的 JSON 数据接口仍是 stocklist.aspx，需在请求头加
    X-Requested-With: XMLHttpRequest 才能触发 JSON 响应。
    已将 STOCKLIST_URL 改回 stocklist.aspx，并在 fetch_stock_list() 中补充该请求头。

  Bug 2 - 持股表格选择器：
    HKEX 页面改版后，结果表格的 class 由单一 "table" 扩充为
    "table table-scroll table-sort table-mobile-list"。
    BeautifulSoup 的 find("table", {"class": "table"}) 进行精确匹配，已无法命中。
    改用 CSS 选择器 soup.select_one("table.table") 进行包含匹配，兼容新旧版本。

  Bug 3 - 连接被服务器强制断开（ConnectionResetError 10054）：
    连续高频请求同一 Session，服务器会主动踢掉长连接。
    修复：每 REFRESH_SESSION_EVERY 次请求刷新一次 Session，
    并在请求失败时自动重试（指数退避，最多 MAX_RETRIES 次）。
"""

import re
import time
from datetime import datetime

import requests
from bs4 import BeautifulSoup
import pandas as pd

# ─────────────────────────────────────────
# 配置
# ─────────────────────────────────────────
DATE_NEW   = "2026/03/27"   # 较新日期（格式 yyyy/mm/dd）
DATE_OLD   = "2026/03/26"   # 较旧日期（格式 yyyy/mm/dd）
THRESHOLD  = 0.30           # 变化率阈值，0.30 = 30%
TOP_N      = 10             # 测试用：只取前 N 只股票，None = 全部

SEARCH_URL   = "https://www3.hkexnews.hk/sdw/search/searchsdw.aspx"

# ── Bug 1 修复 ──────────────────────────────────────────────────────────────
# ccass_stock_list.htm 是纯前端 SPA，直接 GET 只返回 HTML 壳，resp.json() 必然报错。
# 真正的 JSON 数据接口仍是 stocklist.aspx（与旧版相同），故改回此地址。
STOCKLIST_URL = "https://www3.hkexnews.hk/sdw/search/stocklist.aspx"
# ────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": SEARCH_URL,
}

# ── Bug 3 修复：防止服务器强制断连 ─────────────────────────────────────────
REFRESH_SESSION_EVERY = 30    # 每查询 N 只股票后重建一次 Session
MAX_RETRIES           = 4     # 单次请求最大重试次数
RETRY_BASE_DELAY      = 3     # 首次重试等待秒数（指数退避：3→6→12→24）
REQUEST_DELAY         = 0.5   # 每次请求后的基础间隔（秒），比原来略长
# ────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────
# Step 1: 获取 Session / ViewState
# ─────────────────────────────────────────
def get_session():
    session = requests.Session()
    resp = session.get(SEARCH_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    def _val(id_):
        tag = soup.find("input", {"id": id_})
        return tag["value"] if tag else ""

    vs    = _val("__VIEWSTATE")
    vs_gn = _val("__VIEWSTATEGENERATOR")
    print(f"  ✓ Session OK  ViewState={vs[:40]}...")
    return session, vs, vs_gn


# ─────────────────────────────────────────
# Step 2: 获取全量股票列表
# ─────────────────────────────────────────
def fetch_stock_list(date: str, top_n=None) -> list:
    """
    请求 stocklist.aspx?sortby=stockcode&shareholdingdate=YYYYMMDD
    返回 JSON 格式：[{"c": "00001", "n": "CK HUTCHISON..."}, ...]

    关键：必须携带 X-Requested-With: XMLHttpRequest，
    否则服务器返回 HTML 而非 JSON。
    """
    date_compact = date.replace("/", "")   # 20260512
    url = f"{STOCKLIST_URL}?sortby=stockcode&shareholdingdate={date_compact}"

    print(f"\n[Step 2] 获取股票列表: {url}")

    # ── Bug 1 修复：补充 XHR 标志头，使服务器返回 JSON ──
    req_headers = {
        **HEADERS,
        "Referer": "https://www3.hkexnews.hk/sdw/search/ccass_stock_list.htm",
        "X-Requested-With": "XMLHttpRequest",
        "Accept": "application/json, text/javascript, */*; q=0.01",
    }
    resp = requests.get(url, headers=req_headers, timeout=30)
    resp.raise_for_status()

    # 服务器现在正确返回 JSON：[{"c": "00001", "n": "NAME"}, ...]
    data = resp.json()
    stocks = [(item["c"], item["n"]) for item in data if "c" in item and "n" in item]

    print(f"  ✓ 共获取 {len(stocks)} 只股票")

    if top_n:
        stocks = stocks[:top_n]
        print(f"  ► 测试模式：仅取前 {top_n} 只")

    return stocks


# ─────────────────────────────────────────
# Step 3: 查询单只股票在某日各 Broker 持股
# ─────────────────────────────────────────
def _parse_brokers(html: str) -> dict:
    """从持股查询结果 HTML 中解析出 {broker_id: shareholding} 字典。"""
    soup = BeautifulSoup(html, "html.parser")

    alert = soup.find("input", {"id": "alertMsg"})
    if alert and alert.get("value", "").strip():
        return {}

    # Bug 2 修复：用 CSS 选择器包含匹配，兼容 class 多值的新版表格
    table = soup.select_one("table.table")
    if not table:
        return {}

    def cell_val(td):
        body = td.find("div", class_="mobile-list-body")
        return body.get_text(strip=True) if body else td.get_text(strip=True)

    brokers = {}
    for row in table.find_all("tr")[1:]:
        cols = row.find_all("td")
        if len(cols) < 4:
            continue
        pid          = cell_val(cols[0])
        shareholding = cell_val(cols[3]).replace(",", "")
        if not pid.upper().startswith("B"):
            continue
        try:
            brokers[pid] = int(shareholding)
        except ValueError:
            continue

    return brokers


def fetch_broker_holdings(session_state: dict, stock_code: str, date: str) -> dict:
    """
    POST 查询指定股票在某日的经纪商持股。
    session_state: {"session": ..., "vs": ..., "vs_gn": ...} 的可变字典，
                   重试时可在内部刷新，调用方持有同一引用。

    Bug 3 修复：
      - 遇到连接错误自动重试，指数退避（3→6→12→24 秒）
      - 重试时重建 Session + 刷新 ViewState，避免旧连接被服务器拒绝
    """
    payload_base = {
        "__EVENTTARGET":    "btnSearch",
        "__EVENTARGUMENT":  "",
        "txtShareholdingDate": date,
        "txtStockCode":     stock_code,
        "txtStockName":     "",
        "txtParticipantID": "",
        "txtParticipantName": "",
        "sortBy":           "shareholding",
        "sortDirection":    "desc",
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            payload = {
                **payload_base,
                "__VIEWSTATE":          session_state["vs"],
                "__VIEWSTATEGENERATOR": session_state["vs_gn"],
            }
            resp = session_state["session"].post(
                SEARCH_URL, headers=HEADERS, data=payload, timeout=30
            )
            resp.raise_for_status()
            return _parse_brokers(resp.text)

        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout) as e:
            wait = RETRY_BASE_DELAY * (2 ** (attempt - 1))   # 3,6,12,24 秒
            if attempt == MAX_RETRIES:
                print(f"\n    ✗ 重试 {MAX_RETRIES} 次后仍失败，跳过该股票。最后错误: {e}")
                return {}
            print(f"\n    ⚠ 连接中断（{e.__class__.__name__}），"
                  f"{wait}s 后第 {attempt} 次重试（重建 Session）...", end="", flush=True)
            time.sleep(wait)
            # 重建 Session 与 ViewState，避免复用被踢掉的连接
            new_sess, new_vs, new_vsgn = get_session()
            session_state["session"] = new_sess
            session_state["vs"]      = new_vs
            session_state["vs_gn"]   = new_vsgn


# ─────────────────────────────────────────
# Step 4: 生成 xlsx（数据表 + TOP10 条状图）
# ─────────────────────────────────────────
def export_xlsx(df: pd.DataFrame, output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1：完整数据 ──
    ws1 = wb.active
    ws1.title = "持股变化数据"

    headers = list(df.columns)
    hdr_fill   = PatternFill("solid", start_color="1F4E79")
    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    alt_fill   = PatternFill("solid", start_color="EBF3FB")
    border_s   = Side(style="thin", color="CCCCCC")
    cell_bdr   = Border(left=border_s, right=border_s, top=border_s, bottom=border_s)

    for ci, col_name in enumerate(headers, 1):
        c = ws1.cell(row=1, column=ci, value=col_name)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = cell_bdr

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, value in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=value)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center")
            c.border    = cell_bdr
            if ri % 2 == 0:
                c.fill = alt_fill

    for ci, col_name in enumerate(headers, 1):
        max_len = max(len(str(col_name)),
                      df.iloc[:, ci - 1].astype(str).map(len).max())
        ws1.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 35)
    ws1.freeze_panes = "A2"

    # ── Sheet 2：变化率 TOP10 条状图 ──
    ws2 = wb.create_sheet("变化率 TOP 10")

    pct_num = df["最大变化率"].str.replace("%", "").str.replace("+", "").astype(float).abs()
    top10 = df.assign(_pct_num=pct_num).nlargest(10, "_pct_num").sort_values("_pct_num", ascending=False).drop(columns=["_pct_num"])

    ws2["A1"] = "股票代号"
    ws2["B1"] = "最大变化率(%)"
    ws2["A1"].font = Font(name="Arial", bold=True)
    ws2["B1"].font = Font(name="Arial", bold=True)

    for i, (_, row) in enumerate(top10.iterrows(), 2):
        pct_val = float(row["最大变化率"].replace("%", "").replace("+", ""))
        ws2.cell(row=i, column=1, value=row["股票代号"]).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=2, value=pct_val).font        = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16

    chart = BarChart()
    chart.type    = "bar"
    chart.title   = "变化率 TOP 10"
    chart.y_axis.title = "股票代号"
    chart.x_axis.title = "最大变化率(%)"
    chart.style   = 10
    chart.width   = 22
    chart.height  = 14

    chart.x_axis.scaling.logBase = 10
    chart.x_axis.scaling.min     = 1

    data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(top10) + 1)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(top10) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, "D2")

    wb.save(output_path)
    print(f"  ✓ xlsx 已保存至 {output_path}")


# ─────────────────────────────────────────
# 主函数
# ─────────────────────────────────────────
def main():
    print("=" * 65)
    print("CCASS Inter-Broker Movement Analysis")
    print(f"查询页面 : {SEARCH_URL}")
    print(f"对比日期 : {DATE_OLD}  →  {DATE_NEW}")
    print(f"变化率阈值: >{THRESHOLD*100:.0f}%   股票数: {TOP_N if TOP_N else '全部'}")
    print("=" * 65)

    # Step 2: 获取股票列表
    stocks = fetch_stock_list(DATE_NEW, top_n=TOP_N)

    # Step 3: 查询新日期各股票经纪商持股
    print(f"\n[Step 3] 查询 {DATE_NEW} 各股票经纪商持股...")
    sess, vs, vs_gn = get_session()
    state_new = {"session": sess, "vs": vs, "vs_gn": vs_gn}
    holdings_new = {}
    for i, (code, name) in enumerate(stocks, 1):
        # 每隔 REFRESH_SESSION_EVERY 只股票主动刷新 Session，避免长连接被踢
        if i > 1 and (i - 1) % REFRESH_SESSION_EVERY == 0:
            print(f"\n  [Session 刷新] 已查询 {i-1} 只，主动重建 Session...")
            sess, vs, vs_gn = get_session()
            state_new.update({"session": sess, "vs": vs, "vs_gn": vs_gn})

        print(f"  [{i}/{len(stocks)}] {code} {name} @ {DATE_NEW}", end=" ... ", flush=True)
        brokers = fetch_broker_holdings(state_new, code, DATE_NEW)
        holdings_new[code] = brokers
        print(f"{len(brokers)} brokers")
        time.sleep(REQUEST_DELAY)

    # Step 4: 查询旧日期各股票经纪商持股
    print(f"\n[Step 4] 查询 {DATE_OLD} 各股票经纪商持股...")
    sess, vs, vs_gn = get_session()
    state_old = {"session": sess, "vs": vs, "vs_gn": vs_gn}
    holdings_old = {}
    for i, (code, name) in enumerate(stocks, 1):
        if i > 1 and (i - 1) % REFRESH_SESSION_EVERY == 0:
            print(f"\n  [Session 刷新] 已查询 {i-1} 只，主动重建 Session...")
            sess, vs, vs_gn = get_session()
            state_old.update({"session": sess, "vs": vs, "vs_gn": vs_gn})

        print(f"  [{i}/{len(stocks)}] {code} {name} @ {DATE_OLD}", end=" ... ", flush=True)
        brokers = fetch_broker_holdings(state_old, code, DATE_OLD)
        holdings_old[code] = brokers
        print(f"{len(brokers)} brokers")
        time.sleep(REQUEST_DELAY)

    # Step 5: 计算变化，筛选超过阈值的股票
    print("\n[Step 5] 计算经纪商持股变化...")
    records = []

    stock_dict = {code: name for code, name in stocks}

    for code in holdings_new:
        brokers_new = holdings_new.get(code, {})
        brokers_old = holdings_old.get(code, {})

        # 取两日都出现的经纪商
        common_brokers = set(brokers_new.keys()) & set(brokers_old.keys())
        if not common_brokers:
            continue

        max_pct   = 0.0
        max_pid   = ""
        triggered = False

        for pid in common_brokers:
            sh_new = brokers_new[pid]
            sh_old = brokers_old[pid]
            if sh_old == 0:
                continue
            pct = (sh_new - sh_old) / sh_old
            if abs(pct) > abs(max_pct):
                max_pct = pct
                max_pid = pid
            if abs(pct) > THRESHOLD:
                triggered = True

        if triggered:
            records.append({
                "股票代号":   code,
                "股票名称":   stock_dict.get(code, ""),
                "触发经纪商": max_pid,
                f"{DATE_OLD} 持股量": brokers_old.get(max_pid, 0),
                f"{DATE_NEW} 持股量": brokers_new.get(max_pid, 0),
                "最大变化率": f"{max_pct*100:+.2f}%",
            })

    if not records:
        print(f"\n✗ 没有股票的经纪商持股变化超过 {THRESHOLD*100:.0f}%")
        return

    df = pd.DataFrame(records)
    df["_pct"] = df["最大变化率"].str.replace("%", "").str.replace("+", "").astype(float).abs()
    df = df.sort_values("_pct", ascending=False).drop(columns=["_pct"]).reset_index(drop=True)

    print(f"\n── 共 {len(df)} 只股票的经纪商持股变化超过 {THRESHOLD*100:.0f}% ──")
    print(df.to_string(index=False))

    # 保存 CSV
    csv_file = f"ccass_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    df.to_csv(csv_file, index=False, encoding="utf-8-sig")
    print(f"\n✓ CSV 已保存至 {csv_file}")

    # 生成 xlsx
    print("\n[Step 6] 生成 xlsx...")
    xlsx_file = f"ccass_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_xlsx(df.copy(), xlsx_file)


if __name__ == "__main__":
    main()